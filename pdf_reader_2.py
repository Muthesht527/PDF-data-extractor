import PyPDF2
from PyPDF2 import PdfWriter
from pathlib import Path
import pandas as pd
import pytesseract as tess
from pytesseract import Output
import cv2
import re
import shutil
import hashlib
from openpyxl import load_workbook

INPUT_DIR = Path("PDF_Input")
PROCESSED_DIR = Path("Processed_PDFs")
TEMP_DIR = Path("Temp_Img")
CERT_DIR = Path("Certificates")

INPUT_DIR.mkdir(exist_ok=True)
PROCESSED_DIR.mkdir(exist_ok=True)
TEMP_DIR.mkdir(exist_ok=True)
CERT_DIR.mkdir(exist_ok=True)

pdfs = list(INPUT_DIR.glob("*.pdf"))
excel_file = "student_record.xlsx"

def pdf_hash(pdf_path):
    h = hashlib.sha256()
    with open(pdf_path, "rb") as f:
        for chunk in iter(lambda: f.read(8192), b""):
            h.update(chunk)
    return h.hexdigest()

def preprocess_for_ocr(img):
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    gray = cv2.threshold(gray, 150, 255, cv2.THRESH_BINARY)[1]

    try:
        osd = tess.image_to_osd(gray)
        rotation = int(re.search(r"Rotate: (\d+)", osd).group(1))

        if rotation == 90:
            gray = cv2.rotate(gray, cv2.ROTATE_90_CLOCKWISE)
        elif rotation == 180:
            gray = cv2.rotate(gray, cv2.ROTATE_180)
        elif rotation == 270:
            gray = cv2.rotate(gray, cv2.ROTATE_90_COUNTERCLOCKWISE)

    except:
        pass

    return gray

def extract_images(reader, page_index, output_dir):
    page = reader.pages[page_index]
    images = getattr(page, "images", None)
    if not images:
        return []

    saved = []
    for i, image in enumerate(images, start=1):
        ext = getattr(image, "extension", None) or "jpeg"
        path = output_dir / f"page_{page_index+1:03d}_img_{i:03d}.{ext}"
        with open(path, "wb") as f:
            f.write(image.data)
        saved.append(path)
    return saved

def Page1(reader, page_index, output_dir):
    saved = extract_images(reader, page_index, output_dir)
    if not saved:
        raise FileNotFoundError("No image on page 1")

    data = tess.image_to_data(str(saved[0]), output_type=Output.DATAFRAME)
    data["conf"] = pd.to_numeric(data["conf"], errors="coerce")
    data = data[(data.conf > 40) & data.text.notna() & (data.text.str.strip() != "")]

    lines = {}
    for _, r in data.iterrows():
        key = (r.block_num, r.par_num, r.line_num)
        lines.setdefault(key, []).append(r)

    structured = []
    for line in lines.values():
        line = sorted(line, key=lambda r: r.left)
        structured.append({
            "text": " ".join(r.text for r in line),
            "y": min(r.top for r in line)
        })

    # Project title
    report_y = next(l["y"] for l in structured if "project report" in l["text"].lower())
    title_lines = [l for l in structured if l["y"] < report_y]
    title = " ".join(l["text"] for l in title_lines[-3:])

    # Students
    start_y = next(l["y"] for l in structured if "submitted by" in l["text"].lower())
    students = []

    students = []

    for l in structured:
        if l["y"] > start_y and "partial fulfillment" not in l["text"].lower():

            text_line = l["text"].strip()

            # FORMAT 1 → With brackets
            m1 = re.search(r"([A-Za-z\s]+)\s*\((\d{10,15})\)", text_line)

            # FORMAT 2 → Without brackets
            m2 = re.search(r"([A-Za-z\s]+)\s+(\d{10,15})$", text_line)

            match = m1 if m1 else m2

            if match:
                students.append({
                    "Name": match.group(1).strip(),
                    "Register Number": match.group(2),
                    "Project Title": title
                })


    return students

def Page2(reader, page_index, output_dir):
    if page_index >= len(reader.pages):
        return None

    saved = extract_images(reader, page_index, output_dir)
    if not saved:
        return None

    img = cv2.imread(str(saved[0]))
    gray = preprocess_for_ocr(img)

    text = tess.image_to_string(gray, config="--psm 6").lower()
    m = re.search(r"supervisor,\s*(dr\.\s*[a-z.\s]+)", text)
    if m==None:
        m = re.search(r"mentor,\s*(dr\.\s*[a-z.\s]+)", text)
    return m.group(1).strip() if m else None

def is_certificate_page(reader, page_index):
    saved = extract_images(reader, page_index, TEMP_DIR)
    if not saved:
        return False, None, None

    img = cv2.imread(str(saved[0]))
    gray = preprocess_for_ocr(img)

    text = tess.image_to_string(gray, config="--psm 6").lower()

    # ❌ Exclude Bonafide / Academic internal certificate
    if "bonafide certificate" in text:
        return False, None, None

    if "project report" in text:
        return False, None, None

    # =========================
    # JOURNAL PUBLICATION LOGIC
    # =========================
    journal_score = 0

    if "certificate of publication" in text:
        journal_score += 2

    if "issn" in text:
        journal_score += 1

    if "volume" in text and "issue" in text:
        journal_score += 1

    if "published in" in text:
        journal_score += 1

    if journal_score >= 2:
        return True, "journal_publication", text

    # =========================
    # CONFERENCE LOGIC
    # =========================
    conference_score = 0

    if "certificate of participation" in text:
        conference_score += 2

    if "conference" in text:
        conference_score += 1

    if "presented a paper" in text:
        conference_score += 1

    if "organized by" in text:
        conference_score += 1

    if re.search(r"\b\d{4}\b", text):  # year detection
        conference_score += 1

    if conference_score >= 2:
        return True, "conference", text

    return False, None, None

def save_certificate(reader, page_index, filename):
    writer = PdfWriter()
    writer.add_page(reader.pages[page_index])
    with open(filename, "wb") as f:
        writer.write(f)

def extract_certificates(reader, pdf_id, pdf_name, student_names, supervisor, max_pages=15):
    certs = []
    start = max(0, len(reader.pages) - max_pages)

    for i in reversed(range(start, len(reader.pages))):

        ok, cert_category, text = is_certificate_page(reader, i)
        if not ok:
            continue

        cert_role = "unknown"
        cert_name = None

        # 🔹 Match student names
        for s in student_names:
            if s.lower() in text:
                cert_role = "student"
                cert_name = s
                break

        # 🔹 Match supervisor
        if cert_role == "unknown" and supervisor and supervisor.lower() in text:
            cert_role = "supervisor"
            cert_name = supervisor

        safe = cert_name.replace(" ", "_") if cert_name else "unknown"

        fname = CERT_DIR / f"{pdf_id}_{cert_category}_{cert_role}_{safe}_page{i+1}.pdf"
        save_certificate(reader, i, fname)

        certs.append({
            "category": cert_category,
            "role": cert_role,
            "name": cert_name,
            "path": fname.as_posix()
        })

    return certs

def apply_hyperlinks(excel_file, columns):
    wb = load_workbook(excel_file)
    ws = wb.active
    headers = [c.value for c in ws[1]]

    for col in columns:
        if col not in headers:
            continue
        idx = headers.index(col) + 1
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(r, idx)
            if cell.value:
                link = str(cell.value).split(";")[0]
                cell.value = link
                cell.hyperlink = link
                cell.style = "Hyperlink"
    wb.save(excel_file)

def process_single_pdf(pdf_path):
    print(f"Processing: {pdf_path.name}")

    if TEMP_DIR.exists():
        shutil.rmtree(TEMP_DIR)
    TEMP_DIR.mkdir()

    reader = PyPDF2.PdfReader(pdf_path)
    pdf_id = pdf_hash(pdf_path)

    students = Page1(reader, 0, TEMP_DIR)
    supervisor = Page2(reader, 1, TEMP_DIR)

    certs = extract_certificates(
        reader,
        pdf_id,
        pdf_path.name,
        [s["Name"] for s in students],
        supervisor
    )

    df = pd.DataFrame(students)
    df["Register Number"] = df["Register Number"].str.replace(r"\D", "", regex=True)
    df["PDF_ID"] = pdf_id
    df["PDF Name"] = pdf_path.name
    df["PDF Link"] = f"Processed_PDFs/{pdf_path.name}"

    df["Student Certificate Links"] = df["Name"].apply(
        lambda n: "; ".join(
            c["path"] for c in certs
            if c["role"] == "student" and c["name"] == n
        )
    )

    sup_links = "; ".join(
        c["path"] for c in certs
        if c["role"] == "supervisor"
    )

    df["Supervisor"] = supervisor
    df["Supervisor Certificate Links"] = sup_links if sup_links else None

    shutil.move(pdf_path, PROCESSED_DIR / pdf_path.name)
    shutil.rmtree(TEMP_DIR)

    return df

def upload_excel(all_data):
    if not all_data:
        return

    df = pd.concat(all_data, ignore_index=True)

    if Path(excel_file).exists():
        old = pd.read_excel(excel_file)
        df = pd.concat([old, df], ignore_index=True)

    df.drop_duplicates(subset=["PDF_ID", "Register Number"], inplace=True)
    df.to_excel(excel_file, index=False)

    apply_hyperlinks(
        excel_file,
        ["PDF Link", "Student Certificate Links", "Supervisor Certificate Links"]
    )

    print("✅ ALL PDFs PROCESSED")

# ==============================
# MAIN
# ==============================

all_data = []

for pdf in pdfs:
    try:
        all_data.append(process_single_pdf(pdf))
    except Exception as e:
        print(f"❌ Failed: {pdf.name}")
        print(e)

upload_excel(all_data)
